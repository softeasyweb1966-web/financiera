"""
Script de migración de datos desde Excel al nuevo modelo relacional.
Lee 'GASTOS FIJOS MENSUALES 2026 .xlsx' y crea terceros, servicios,
obligaciones, empleados y sus pagos correspondientes.

Uso: venv\Scripts\python.exe importar_excel.py
"""
import openpyxl
from app import create_app, db
from app.models import (
    Tercero, TipoTercero, Categoria, Concepto,
    Servicio, PagoServicio, Obligacion, PagoObligacion,
    Empleado, RegistroNomina, ConceptoNomina
)

ANIO = 2026
EXCEL_FILE = 'GASTOS FIJOS MENSUALES 2026 .xlsx'

# Mapeo de empresas del Excel a conceptos de servicio
CONCEPTO_SERVICIO_MAP = {
    'BIOFILE': 'Software/Plataforma',
    'CELULARES': 'Celular',
    'ATICA': 'Software/Plataforma',
    'ARRIENDO': 'Arriendo',
    'ACUEDUCTO 1 PISO': 'Acueducto',
    'ACUEDUCTO2 PISO': 'Acueducto',
    'PROEXEQUIAL': 'Plan exequial',
    'CLARO': 'Teléfono',
    'CODENSA 2 PISO': 'Energía',
    'CODENSA PREVENTSALUD': 'Energía',
    'ETB': 'Internet',
    'GAS 2 PISO': 'Gas',
    'CELULAR JUAN': 'Celular',
    'CELULAR CONTABILIDAD': 'Celular',
    'CELULAR SEBASTIAN COMERCIAL': 'Celular',
    'CELULAR ASISTENTE COMERCIAL': 'Celular',
    'SIIGO': 'Software/Plataforma',
    'SUPERSALUD': 'Otro Servicio',
    'DIAN': 'Otro Servicio',
}

# Mapeo de obligaciones a modalidades
MODALIDAD_MAP = {
    'AV VILLAS': 'bancario_cuota_fija',
    'COLPATRIA': 'bancario_cuota_fija',
    'CIELO': 'bancario_cuota_fija',
    'KAREN': 'solo_interes',
    'PRESTAMO MERCEDEZ': 'solo_interes',
    'CADENA': 'cadena',
    'OSCAR DUSAN': 'pago_total_pactado',
}

# Mapeo de obligaciones a conceptos
CONCEPTO_OBLIGACION_MAP = {
    'AV VILLAS': 'Cuota hipotecaria',
    'COLPATRIA': 'Cuota hipotecaria',
    'CIELO': 'Cuota consumo',
    'KAREN': 'Interés préstamo personal',
    'PRESTAMO MERCEDEZ': 'Interés préstamo personal',
    'CADENA': 'Cadena',
    'OSCAR DUSAN': 'Interés préstamo personal',
}

stats = {'terceros': 0, 'servicios': 0, 'pagos_servicios': 0,
         'obligaciones': 0, 'pagos_obligaciones': 0,
         'empleados': 0, 'pagos_nomina': 0, 'omitidos': 0, 'errores': 0}


def parse_valor(val):
    """Convierte valor a float o None"""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip().upper().replace('$', '').replace('.', '').replace(',', '')
    if val_str in ('N.A', 'NA', 'N.A.', 'U', 'SE RECIBE', '#NAME?', ''):
        return None
    try:
        return float(val_str)
    except ValueError:
        return None


def estado_from_valor(val_raw):
    """Determina estado basado en valor crudo"""
    if val_raw is None or val_raw == '':
        return 'pendiente'
    if isinstance(val_raw, str) and val_raw.strip().upper() in ('N.A', 'NA', 'N.A.'):
        return 'n/a'
    return 'pagado'


def get_or_create_tercero(nombre, tipo_nombre):
    """Busca o crea un tercero"""
    nombre = nombre.strip().upper()
    tipo = TipoTercero.query.filter_by(nombre=tipo_nombre).first()
    tercero = Tercero.query.filter_by(nombre=nombre, tipo_tercero_id=tipo.id).first()
    if not tercero:
        tercero = Tercero(nombre=nombre, tipo_tercero_id=tipo.id)
        db.session.add(tercero)
        db.session.flush()
        stats['terceros'] += 1
    return tercero


def get_concepto(nombre, categoria_nombre):
    """Busca un concepto por nombre y categoría"""
    cat = Categoria.query.filter_by(nombre=categoria_nombre).first()
    concepto = Concepto.query.filter_by(nombre=nombre, categoria_id=cat.id).first()
    if not concepto:
        concepto = Concepto(nombre=nombre, categoria_id=cat.id)
        db.session.add(concepto)
        db.session.flush()
    return concepto


def importar_servicios():
    print('\n=== IMPORTANDO SERVICIOS ===')
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['SERVICIOS']

    # Detectar columnas de meses (fila 2)
    headers = [cell.value for cell in ws[2]]
    MESES_MAP = {
        'ENERO': 1, 'FEBRERO': 2, 'FEBRERO ': 2, 'MARZO': 3,
        'ABRIL': 4, 'ABRIL ': 4, 'MAYO': 5, 'MAYO ': 5,
        'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9,
        'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12
    }
    meses_cols = []
    for i, h in enumerate(headers):
        if h and str(h).strip().upper() in MESES_MAP:
            meses_cols.append((i, MESES_MAP[str(h).strip().upper()]))

    for row in ws.iter_rows(min_row=3, max_row=22, values_only=True):
        empresa_raw = row[1]
        if not empresa_raw or str(empresa_raw).strip().upper() in ('TOTALES', 'DIAN'):
            continue

        empresa = str(empresa_raw).strip().upper()
        dia_pago = str(row[2]).strip() if row[2] else None
        referencia = str(row[3]).strip() if row[3] else None
        valor_estimado = parse_valor(row[4])

        # Determinar concepto
        concepto_nombre = CONCEPTO_SERVICIO_MAP.get(empresa, 'Otro Servicio')
        concepto = get_concepto(concepto_nombre, 'Servicios Públicos')

        # Crear tercero (empresa de servicios)
        tercero = get_or_create_tercero(empresa, 'Empresa Servicios')

        # Crear servicio
        servicio = Servicio.query.filter_by(tercero_id=tercero.id, referencia=referencia).first()
        if not servicio:
            servicio = Servicio(
                tercero_id=tercero.id,
                concepto_id=concepto.id,
                referencia=referencia if referencia and referencia != 'None' else None,
                dia_limite_pago=int(dia_pago) if dia_pago and dia_pago.isdigit() else None,
                periodicidad='anual' if dia_pago and dia_pago.upper() == 'ANUAL' else 'mensual',
                valor_estimado=valor_estimado
            )
            db.session.add(servicio)
            db.session.flush()
            stats['servicios'] += 1

        # Registrar pagos mensuales
        for col_idx, mes_num in meses_cols:
            if col_idx < len(row):
                val_raw = row[col_idx]
                estado = estado_from_valor(val_raw)
                valor = parse_valor(val_raw)

                if estado != 'pendiente':
                    existe = PagoServicio.query.filter_by(
                        servicio_id=servicio.id, anio=ANIO, mes=mes_num
                    ).first()
                    if not existe:
                        pago = PagoServicio(
                            servicio_id=servicio.id, anio=ANIO, mes=mes_num,
                            valor_pagado=valor, estado=estado
                        )
                        db.session.add(pago)
                        stats['pagos_servicios'] += 1
                    else:
                        stats['omitidos'] += 1

    db.session.commit()
    print(f'  Terceros creados: {stats["terceros"]}')
    print(f'  Servicios creados: {stats["servicios"]}')
    print(f'  Pagos registrados: {stats["pagos_servicios"]}')


def importar_bancos():
    print('\n=== IMPORTANDO OBLIGACIONES BANCARIAS ===')
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['BANCOS']

    headers = [cell.value for cell in ws[1]]
    MESES_MAP = {
        'ENERO': 1, 'FEBRERO': 2, 'FEBRERO ': 2, 'MARZO': 3,
        'ABRIL': 4, 'ABRIL ': 4, 'MAYO': 5, 'MAYO ': 5,
        'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8
    }
    meses_cols = []
    for i, h in enumerate(headers):
        if h and str(h).strip().upper() in MESES_MAP:
            meses_cols.append((i, MESES_MAP[str(h).strip().upper()]))

    for row in ws.iter_rows(min_row=2, max_row=9, values_only=True):
        banco_raw = row[0]
        if not banco_raw or str(banco_raw).strip().upper() == 'TOTAL':
            continue

        banco = str(banco_raw).strip().upper()
        valor_capital = parse_valor(row[1])
        fecha_limite = int(row[2]) if row[2] and str(row[2]).isdigit() else None
        obligacion_nombre = str(row[3]).strip() if row[3] else banco
        titular = str(row[4]).strip() if row[4] else None

        # Determinar tipo de tercero y modalidad
        modalidad = MODALIDAD_MAP.get(banco, 'bancario_cuota_fija')
        concepto_nombre = CONCEPTO_OBLIGACION_MAP.get(banco, 'Otro bancario')

        if modalidad in ('bancario_cuota_fija', 'cadena'):
            tipo_tercero = 'Entidad Financiera'
        else:
            tipo_tercero = 'Prestamista Personal'

        tercero = get_or_create_tercero(banco, tipo_tercero)
        concepto = get_concepto(concepto_nombre, 'Obligaciones Bancarias')

        # Crear obligación
        obligacion = Obligacion.query.filter_by(
            tercero_id=tercero.id, titular=titular
        ).first()
        if not obligacion:
            obligacion = Obligacion(
                tercero_id=tercero.id,
                concepto_id=concepto.id,
                modalidad=modalidad,
                capital_inicial=valor_capital,
                saldo_actual=valor_capital,
                dia_limite_pago=fecha_limite,
                titular=titular
            )
            db.session.add(obligacion)
            db.session.flush()
            stats['obligaciones'] += 1

        # Registrar pagos
        for col_idx, mes_num in meses_cols:
            if col_idx < len(row):
                val_raw = row[col_idx]
                estado = estado_from_valor(val_raw)
                valor = parse_valor(val_raw)

                if estado != 'pendiente':
                    existe = PagoObligacion.query.filter_by(
                        obligacion_id=obligacion.id, anio=ANIO, mes=mes_num
                    ).first()
                    if not existe:
                        pago = PagoObligacion(
                            obligacion_id=obligacion.id, anio=ANIO, mes=mes_num,
                            valor_pagado=valor, estado=estado
                        )
                        db.session.add(pago)
                        stats['pagos_obligaciones'] += 1
                    else:
                        stats['omitidos'] += 1

    db.session.commit()
    print(f'  Obligaciones creadas: {stats["obligaciones"]}')
    print(f'  Pagos registrados: {stats["pagos_obligaciones"]}')


def importar_nomina():
    print('\n=== IMPORTANDO NÓMINA ===')
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['NOMINA']

    # Concepto Salario por defecto
    concepto_salario = ConceptoNomina.query.filter_by(nombre='Salario').first()
    concepto_honorarios = ConceptoNomina.query.filter_by(nombre='Honorarios').first()
    concepto_parafiscales = ConceptoNomina.query.filter_by(nombre='Parafiscales').first()
    concepto_seg_social = ConceptoNomina.query.filter_by(nombre='Seguridad Social').first()

    # Columnas: 0=NOMBRE, 1=CARGO, 2=SALARIO, 3=ENE-Q1, 4=ENE-Q2, 5=FEB-Q1, 6=FEB-Q2...
    meses_inicio = 3

    # Filas especiales
    FILAS_PARAFISCALES = ('SEGURIDAD SOCIAL', 'PLAN COMPLEMENTARIO DRA ANA')

    for row in ws.iter_rows(min_row=3, max_row=24, values_only=True):
        nombre_raw = row[0]
        if not nombre_raw:
            continue

        nombre = str(nombre_raw).strip().upper()
        cargo = str(row[1]).strip() if row[1] else None
        salario = parse_valor(row[2])

        # Determinar si es parafiscal o empleado regular
        es_parafiscal = nombre in FILAS_PARAFISCALES
        es_prestacion_servicios = cargo and cargo.upper() in (
            'CONTADORA', 'BACTERIOLOGA', 'OPTOMETRA', 'FONOAUDIOLOGA',
            'COMUNITY MANAGER', 'COMUNITY MANAGER '
        )

        if es_parafiscal:
            # Registrar como parafiscal global (sin empleado)
            concepto = concepto_parafiscales if 'SEGURIDAD' not in nombre else concepto_seg_social
            for mes_idx in range(7):
                mes_num = mes_idx + 1
                col_q1 = meses_inicio + (mes_idx * 2)
                col_q2 = col_q1 + 1

                for q, col in [(1, col_q1), (2, col_q2)]:
                    if col < len(row):
                        val_raw = row[col]
                        estado = estado_from_valor(val_raw)
                        valor = parse_valor(val_raw)
                        if estado == 'pagado' and valor:
                            existe = RegistroNomina.query.filter_by(
                                empleado_id=None, concepto_nomina_id=concepto.id,
                                anio=ANIO, mes=mes_num, quincena=q
                            ).first()
                            if not existe:
                                reg = RegistroNomina(
                                    empleado_id=None, concepto_nomina_id=concepto.id,
                                    anio=ANIO, mes=mes_num, quincena=q, valor=valor
                                )
                                db.session.add(reg)
                                stats['pagos_nomina'] += 1
            continue

        # Crear tercero y empleado
        tercero = get_or_create_tercero(nombre, 'Empleado')
        empleado = Empleado.query.filter_by(tercero_id=tercero.id).first()
        if not empleado:
            tipo_contrato = 'prestacion_servicios' if es_prestacion_servicios else 'laboral'
            empleado = Empleado(
                tercero_id=tercero.id,
                cargo=cargo,
                salario_base=salario,
                tipo_contrato=tipo_contrato
            )
            db.session.add(empleado)
            db.session.flush()
            stats['empleados'] += 1

        # Concepto según tipo
        concepto = concepto_honorarios if es_prestacion_servicios else concepto_salario

        # Registrar pagos quincenales
        for mes_idx in range(7):
            mes_num = mes_idx + 1
            col_q1 = meses_inicio + (mes_idx * 2)
            col_q2 = col_q1 + 1

            for q, col in [(1, col_q1), (2, col_q2)]:
                if col < len(row):
                    val_raw = row[col]
                    estado = estado_from_valor(val_raw)
                    valor = parse_valor(val_raw)
                    if estado == 'pagado' and valor:
                        existe = RegistroNomina.query.filter_by(
                            empleado_id=empleado.id, concepto_nomina_id=concepto.id,
                            anio=ANIO, mes=mes_num, quincena=q
                        ).first()
                        if not existe:
                            reg = RegistroNomina(
                                empleado_id=empleado.id, concepto_nomina_id=concepto.id,
                                anio=ANIO, mes=mes_num, quincena=q, valor=valor
                            )
                            db.session.add(reg)
                            stats['pagos_nomina'] += 1
                        else:
                            stats['omitidos'] += 1

    db.session.commit()
    print(f'  Empleados creados: {stats["empleados"]}')
    print(f'  Registros nómina: {stats["pagos_nomina"]}')


def main():
    app = create_app()
    with app.app_context():
        print('=' * 50)
        print(' MIGRACIÓN DE DATOS DESDE EXCEL')
        print(f' Archivo: {EXCEL_FILE}')
        print(f' Año: {ANIO}')
        print('=' * 50)

        importar_servicios()
        importar_bancos()
        importar_nomina()

        print('\n' + '=' * 50)
        print(' REPORTE DE MIGRACIÓN')
        print('=' * 50)
        print(f'  Terceros creados:        {stats["terceros"]}')
        print(f'  Servicios creados:       {stats["servicios"]}')
        print(f'  Pagos servicios:         {stats["pagos_servicios"]}')
        print(f'  Obligaciones creadas:    {stats["obligaciones"]}')
        print(f'  Pagos obligaciones:      {stats["pagos_obligaciones"]}')
        print(f'  Empleados creados:       {stats["empleados"]}')
        print(f'  Registros nómina:        {stats["pagos_nomina"]}')
        print(f'  Omitidos (duplicados):   {stats["omitidos"]}')
        print(f'  Errores:                 {stats["errores"]}')
        total = (stats["pagos_servicios"] + stats["pagos_obligaciones"] + stats["pagos_nomina"])
        print(f'  TOTAL REGISTROS:         {total}')
        print('=' * 50)
        print('\n✓ Migración completada.')
        print('  Inicie la app con: iniciar.bat')


if __name__ == '__main__':
    main()
